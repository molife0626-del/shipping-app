import streamlit as st
import pandas as pd
import io
import openpyxl
from datetime import datetime

# ページ設定
st.set_page_config(page_title="出荷重量計算システム(図番照合版)", layout="wide")

# ==========================================
# 🔐 パスワード認証 (Enterキー対応)
# ==========================================
def check_password():
    SECRET_PASSWORD = "1234" 
    if "password_correct" not in st.session_state:
        st.session_state.password_correct = False

    if not st.session_state.password_correct:
        with st.form("login_form"):
            st.title("🔒 ログイン")
            pwd = st.text_input("パスワード", type="password")
            submitted = st.form_submit_button("ログイン")
            if submitted:
                if pwd == SECRET_PASSWORD:
                    st.session_state.password_correct = True
                    st.rerun()
                else:
                    st.error("パスワードが違います")
        st.stop()

check_password()

# ==========================================
# 🚛 メインアプリ
# ==========================================
st.title("🚛 出荷重量計算システム (図番・品目番号 照合版)")

if 'master_df' not in st.session_state:
    st.session_state.master_df = None

# ---------------------------------------------------------
# STEP 1: マスター登録
# ---------------------------------------------------------
st.header("❶ 単重マスターの登録")
st.markdown("「図番」と「単重」が書かれたファイルを登録します。")
master_file = st.file_uploader("単重マスター(Excel/CSV)", type=None, key="m")

if master_file:
    try:
        if master_file.name.endswith('.csv'): df_m = pd.read_csv(master_file)
        else: df_m = pd.read_excel(master_file)
        
        st.dataframe(df_m.head(3))
        cols_m = df_m.columns.tolist()
        
        c1,c2,c3 = st.columns(3)
        def_key, def_w = 0, 0
        
        # ★自動選択ロジック（図番を優先）
        for i,c in enumerate(cols_m):
            if "図番" in str(c) or "品番" in str(c): def_key=i
            if "(Kg)" in str(c) or "単重" in str(c): def_w=i
            
        m_key = c1.selectbox("照合キー (例: 図番)", cols_m, index=def_key, help="出荷データの「品目番号」と突き合わせる列です")
        m_w = c2.selectbox("重量列 (例: 単重)", cols_m, index=def_w)
        unit = c3.radio("単位", ["kg", "g"], index=0)
        
        if st.button("登録"):
            # 必要な列だけ抽出してリネーム
            # キー列を「MasterKey」、重量を「単重」とする
            cm = df_m[[m_key, m_w]].copy()
            cm.columns = ["MasterKey","単重"]
            
            # 数値変換
            cm["単重"] = pd.to_numeric(cm["単重"], errors='coerce').fillna(0)
            if unit=="g": cm["単重"]=cm["単重"]/1000.0
            
            # 図番が重複している場合は最初の行を採用
            st.session_state.master_df = cm.drop_duplicates("MasterKey")
            st.success(f"登録完了 (キー: {m_key})")
            
    except Exception as e: st.error("エラー: ファイル形式を確認してください")
st.divider()

# ---------------------------------------------------------
# STEP 2: 計算
# ---------------------------------------------------------
st.header("❷ 出荷指示計算")

if st.session_state.master_df is None:
    st.info("先にSTEP 1でマスターを登録してください")
    st.stop()

st.markdown("出荷指示書（注文一覧）をアップロードしてください。")
ship_file = st.file_uploader("出荷指示(Excel/CSV)", type=None, key="s")

if ship_file:
    try:
        if ship_file.name.endswith('.csv'): df_s = pd.read_csv(ship_file)
        else: df_s = pd.read_excel(ship_file)
        
        st.write("▼ 読み込んだデータ")
        st.dataframe(df_s.head(3))
        cols_s = df_s.columns.tolist()
        
        c1,c2,c3,c4 = st.columns(4)
        def_skey, def_sq, def_sname = 0,0,0
        
        # ★自動選択ロジック（品目番号と発注残数を優先）
        for i,c in enumerate(cols_s):
            if "品目番号" in str(c) or "図番" in str(c): def_skey=i
            if "発注残数" in str(c) or "残数" in str(c): def_sq=i
            if "品名" in str(c): def_sname=i
            
        s_key = c1.selectbox("照合キー (例: 品目番号)", cols_s, index=def_skey)
        s_q = c2.selectbox("数量列 (例: 発注残数)", cols_s, index=def_sq)
        s_name = c3.selectbox("品名列 (表示用)", cols_s, index=def_sname)
        max_w = c4.number_input("1パレット上限(kg)", value=500, step=50)

        if st.button("計算実行", type="primary"):
            # 1. 照合（マージ）
            # 出荷データの「品目番号」と マスターの「図番(MasterKey)」を結合
            merged = pd.merge(df_s, st.session_state.master_df, left_on=s_key, right_on="MasterKey", how="left")
            
            # 2. 計算
            merged[s_q] = pd.to_numeric(merged[s_q], errors='coerce').fillna(0)
            merged["総重量"] = merged[s_q] * merged["単重"]
            
            # 3. パレット割付
            pid, cur_w, alloc = 1, 0, []
            for _,r in merged.iterrows():
                w = r["総重量"]
                # 重量0（マスター無しなど）はスキップせずパレットに載せる
                if w > max_w: 
                    alloc.append("超過")
                    continue
                
                if cur_w + w <= max_w:
                    alloc.append(pid)
                    cur_w += w
                else:
                    pid += 1
                    alloc.append(pid)
                    cur_w = w
            
            merged["パレットNo"] = alloc
            
            # 結果用データフレーム作成
            # 表示したい列: パレットNo, 品目番号, 品名, 発注残数, 単重, 総重量
            st.session_state.res_df = merged[["パレットNo", s_key, s_name, s_q, "単重", "総重量"]].rename(
                columns={s_key:"品番(図番)", s_name:"品名", s_q:"数量"}
            )
            
            # 集計
            valid = st.session_state.res_df[pd.to_numeric(st.session_state.res_df["パレットNo"], errors='coerce').notnull()]
            st.session_state.summary = valid.groupby("パレットNo")["総重量"].sum().reset_index()
            
            # 未登録チェック
            unknown = st.session_state.res_df[st.session_state.res_df["単重"].isna() | (st.session_state.res_df["単重"]==0)]
            if not unknown.empty:
                st.warning(f"⚠️ マスターに登録がない（または単重0）の製品が {len(unknown)} 件あります。重量0kgとして計算しています。")
                st.dataframe(unknown)

            st.success("計算完了！")
            
            c_r1, c_r2 = st.columns([1,2])
            c_r1.write("📊 パレット集計")
            c_r1.dataframe(st.session_state.summary, hide_index=True)
            c_r2.write("📋 詳細リスト")
            c_r2.dataframe(st.session_state.res_df, hide_index=True)

    except Exception as e: st.error(f"エラー: {e}")

# --- STEP 3: 印刷 ---
st.divider()
st.header("❸ 印刷用ファイルの出力")

if 'summary' in st.session_state:
    st.markdown("テンプレートExcelをアップロードしてください。")
    template_file = st.file_uploader("テンプレートExcelを選択", type=None, key="tpl")

    if template_file:
        try:
            wb = openpyxl.load_workbook(template_file)
            ws = wb.active

            now = datetime.now()
            if ws['C2'].value is None: ws['C2'] = now.year
            if ws['E2'].value is None: ws['E2'] = now.month
            if ws['G2'].value is None: ws['G2'] = now.day

            summary_dict = dict(zip(st.session_state.summary["パレットNo"], st.session_state.summary["総重量"]))

            # 画像(受領証)の位置に合わせて出力
            ws['H5'] = summary_dict.get(1, 0)
            ws['H6'] = summary_dict.get(2, 0)
            ws['H7'] = summary_dict.get(3, 0)
            ws['H8'] = summary_dict.get(4, 0)
            ws['H16'] = summary_dict.get(5, 0)
            ws['H12'] = summary_dict.get(6, 0)
            
            total_w = st.session_state.summary["総重量"].sum()
            ws['H20'] = total_w

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            st.download_button(
                label="📄 受領証をダウンロード",
                data=output,
                file_name=f"受領証_{now.strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
            
        except Exception as e:
            st.error(f"テンプレートエラー: {e}")
else:
    st.info("計算後に印刷メニューが表示されます。")
